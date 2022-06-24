# 사업자 등록상태 조회 자동화

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time

from bs4 import BeautifulSoup

from openpyxl import load_workbook

wb = load_workbook("./RPAbasic/crawl/RPAselenium/18_business_number.xlsx")
ws = wb.active

browser = webdriver.Chrome()
browser.maximize_window()
browser.get(
    "https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/ab/a/a/UTEABAAA13.xml"
)

# 엑셀 내용 화면 출력
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value)
        if y == 1:
            browser.find_element(By.ID, "bsno").send_keys(ws.cell(x, y).value)
            browser.find_element(By.ID, "trigger5").click()
            time.sleep(1)
            soup = BeautifulSoup(browser.page_source, "lxml")
            td = soup.select("tbody > tr > td")
        elif y == 2:
            ws.cell(row=x, column=y, value=td[1].get_text())
        else:
            ws.cell(row=x, column=y, value=td[2].get_text())

    del soup
time.sleep(3)
browser.quit()

wb.save("./RPAbasic/crawl/RPAselenium/18_business_number.xlsx")
