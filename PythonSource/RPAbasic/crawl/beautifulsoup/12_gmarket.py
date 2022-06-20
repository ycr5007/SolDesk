# Gmarket best - 컴퓨터 / 전자 항목 추출
# base URL : http://corners.gmarket.co.kr/Bestsellers?viewType=G&groupCode=G06

from ctypes import alignment
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# import urllib.request as req
# from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

wb = Workbook()
ws = wb.active

ws.title = "G-Market Best 100"

ws.append(["순위", "제품명", "판매사", "원가", "할인금액", "사진"])

ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 60
ws.column_dimensions["G"].width = 40
# for i in range(2, 102):
#     ws.row_dimensions[i].height = 235

res = requests.get("http://corners.gmarket.co.kr/Bestsellers?viewType=G&groupCode=G06")
soup = BeautifulSoup(res.text, "lxml")

best = soup.select(".best-list > ul > li")


for idx, item in enumerate(best, 1):
    itemname = item.select_one(".itemname")
    item_o_price = item.select_one(".o-price")
    item_s_price = item.select_one(".s-price > strong > span")
    itemimg = item.select_one("img").get("data-original")

    itemsrc = itemname.get("href")
    item_res = requests.get(itemsrc)
    item_soup = BeautifulSoup(item_res.text, "lxml")

    shop = item_soup.select_one("span.text")
    # imgname = f"./RPAbasic/crawl/beautifulsoup/download/{idx}_{shop.get_text()}.jfif"
    # try:
    #     pass
    #     img, header = req.urlretrieve(
    #         ("http:" + itemimg),
    #         imgname,
    #     )
    # except Exception as e:
    #     print(e)
    # else:
    #     print(header)
    # img = Image(imgname)
    ws.append(
        [
            idx,
            itemname.get_text(),
            shop.get_text(),
            item_o_price.get_text(),
            item_s_price.get_text(),
            itemsrc,
        ]
    )
    ws.cell(row=idx + 1, column=6).hyperlink = itemsrc
    # ws.add_image(img, "F" + str(idx + 1))

# 서식 지정
font = Font(name="Tahoma", size=14, color="01579b")
alignment = Alignment(horizontal="center")

cell_a1 = ws["A1"]
cell_a1.alignment = alignment
cell_a1.font = font
cell_b1 = ws["B1"]
cell_b1.alignment = alignment
cell_b1.font = font
cell_c1 = ws["C1"]
cell_c1.alignment = alignment
cell_c1.font = font
cell_d1 = ws["D1"]
cell_d1.alignment = alignment
cell_d1.font = font
cell_e1 = ws["E1"]
cell_e1.alignment = alignment
cell_e1.font = font
cell_f1 = ws["F1"]
cell_f1.alignment = alignment
cell_f1.font = font
cell_g1 = ws["G1"]
cell_g1.alignment = alignment
cell_g1.font = font

wb.save("./RPAbasic/crawl/beautifulsoup/download/G-Market_Best_100.xlsx")
