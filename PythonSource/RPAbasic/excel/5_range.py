from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active

# 한 줄씩 데이터 삽입 - List 이용
ws.append(["번호", "영어", "수학"])  # 첫 줄 입력
for i in range(1, 11):
    ws.append([i, random.randint(0, 100), random.randint(0, 100)])
# ws.append(["", "=average(B2:B11)", "=average(C2:C11)"])

# rows = [
#     ["이름", "생년월일"],
#     ["홍길동", "801010"],
#     ["이몽룡", "000101"],
#     ["성춘향", "630409"],
#     ["조나단", "021212"],
# ]
# for row in rows:
#     ws.append(row)

# 첫번째 행 가져오기
# for y in range(1, ws.max_column + 1):
#     print(ws.cell(1, y).value, end=" ")

first_row = ws[1]  # Tuple 형태로 행에 대한 데이터 반환
for cell in first_row:
    print(cell.value, end=" ")

print()
range_row = ws[2:6]  # 슬라이싱, 마지막 번호까지 포함
for row in range_row:
    for cell in row:
        print(cell.value, end="\t")
    print()

print()
# 3번 행부터 마지막까지 가져오기
for row in ws[3 : ws.max_row]:
    for cell in row:
        print(cell.value, end="\t")
    print()

print()
# iter_rows() / iter_cols()
for row in ws.iter_rows():
    print(row[2].value)

for col in ws.iter_cols():
    print(col[2].value)

wb.save("./RPAbasic/excel/range.xlsx")
