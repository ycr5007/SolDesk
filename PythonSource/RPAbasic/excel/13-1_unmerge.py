# 셀 병합 해제
from openpyxl import load_workbook

wb = load_workbook("./RPAbasic/excel/merge.xlsx")
ws = wb.active

# 병합 셀 범위 지정
ws.unmerge_cells("B2:D2")
ws["B2"].value = "unMerged cell"

wb.save("./RPAbasic/excel/unmerge.xlsx")
