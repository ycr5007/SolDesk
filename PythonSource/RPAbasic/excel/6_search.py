from openpyxl import load_workbook

wb = load_workbook("./RPAbasic/excel/range.xlsx")
ws = wb.active


# 제목행 제거 : min_row > 시작 행 설정

# for row in ws.iter_rows(min_row=2):
#     if row[1].value > 80:
#         print(row[0].value, row[1].value, row[2].value, sep="\t")

# 셀 안에 영어라는 문자가 입력된 셀이 있는 경우 컴퓨터 변경
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("./RPAbasic/excel/range.xlsx")
