from openpyxl import Workbook

wb = Workbook()

# 기본 시트 활성화
ws = wb.active
ws.title = "test"

# 셀에 데이터 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["A4"] = "=sum(A1:A3)"

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6
ws["B4"] = "=B1 + B2 * B3"

# 셀 값 가져오기
print("A1", ws["A1"])  # <Cell 'test'.A1>
print("A1", ws["A1"].value)  # 1
print(ws.cell(1, 1).value)  # 2 차원 좌표형태로 가져올 수 있음 ( x 축 : Row , y 축 : Column )
print(ws.cell(1, 2).value)

# .cell(row, col, value)
c = ws.cell(1, 3, 10)
c = ws.cell(row=1, column=3, value=10)
print(c.value)

wb.save("./RPAbasic/excel/sample.xlsx")
