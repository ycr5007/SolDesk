# Excel 파일 생성

from openpyxl import Workbook

# 1 ) 새 워크북 생성
wb = Workbook()

# 2 ) 현재 활성화 된 시트 가져오기
ws = wb.active

# 3 ) 워크시트 이름 변경
ws.title = "test"

ws["A1"] = 42

ws.append([1, 2, 3])

import datetime

ws["A2"] = datetime.datetime.now()

# 워크북 저장 ( 가상환경을 기준으로 경로를 잡고 있음 )
wb.save("./RPAbasic/excel/sample.xlsx")
