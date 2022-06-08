from openpyxl import Workbook

wb = Workbook()  # Workbook 생성 시 기본 Sheet 생성

# 새로운 시트 생성
ws1 = wb.create_sheet()
ws1.title = "MySheet"
ws1.sheet_properties.tabColor = "ff66ff"

ws2 = wb.create_sheet("SecondSheet", 0)

# 생성된 시트의 모든 이름 출력
print("생성된 Sheet : ", wb.sheetnames)
# 기존 시트 접근
new_ws = wb["MySheet"]
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied MySheet"

wb.save("./RPAbasic/excel/sample.xlsx")
