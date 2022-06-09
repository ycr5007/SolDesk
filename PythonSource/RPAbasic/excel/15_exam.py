import enum
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.title = "성적표"

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])

data_rows = [
    [1, 10, 8, 5, 14, 26, 12],
    [2, 7, 3, 7, 15, 24, 18],
    [3, 9, 5, 8, 8, 12, 4],
    [4, 7, 8, 7, 17, 21, 18],
    [5, 7, 8, 7, 16, 25, 15],
    [6, 3, 5, 8, 8, 17, 0],
    [7, 4, 9, 10, 16, 27, 18],
    [8, 6, 6, 6, 15, 19, 17],
    [9, 10, 10, 9, 19, 30, 19],
]

for row in data_rows:
    ws.append(row)

for x in range(2, ws.max_row + 1):
    ws.cell(x, 4, 10)

ws["H1"].value = "총점"
ws["I1"].value = "성적"

for idx, score in enumerate(data_rows, start=2):
    total = sum(score[1:]) - score[3] + 10
    ws.cell(row=idx, column=8).value = "=sum(B{0}:G{0})".format(idx)
    grade = None
    if total >= 90:
        grade = "A"
    elif total >= 80:
        grade = "B"
    elif total >= 70:
        grade = "C"
    else:
        grade = "D"
    if ws.cell(row=idx, column=2).value < 5:
        grade = "F"
    ws.cell(row=idx, column=9, value=grade)

wb.save("./RPAbasic/excel/scores.xlsx")
